from test.services.fourlog_Listener import create_delivery
from test.services.cache_services import get_delivery_details
from test.services.delivery_generic import CreateDelivery
from convert import Excel_Handler
from utils import build_text_menu
from utils import errhandler
from openpyxl import Workbook, load_workbook
from datetime import datetime
from alive_progress import alive_bar
from pathlib import Path
import sys
import os
import json
import shutil
import tabulate
import csv
import logging
import requests
import PyPDF2

def save_pdffile(location):
    #Sets the scripts working directory to the location of the PDFs
    os.chdir(location)
    #Ask user for the name to save the file as
    userfilename='Output'
    #Get all the PDF filenames
    pdf2merge = []
    for filename in os.listdir('.'):
        if filename.endswith('.pdf'):
            pdf2merge.append(filename)
    pdfWriter = PyPDF2.PdfFileWriter()
    #loop through all PDFs
    for filename in pdf2merge:
        #rb for read binary
        pdfFileObj = open(filename,'rb')
        pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
        #Opening each page of the PDF
        for pageNum in range(pdfReader.numPages):
            pageObj = pdfReader.getPage(pageNum)
            pdfWriter.addPage(pageObj)
    #save PDF to file, wb for write binary
    pdfOutput = open(userfilename+'.pdf', 'wb')
    #Outputting the PDF
    pdfWriter.write(pdfOutput)
    #Closing the PDF writer
    pdfOutput.close()
    #Change Dir to Main Path
    os.chdir(Path(os.path.dirname(__file__)))

def log(error):
    logging.basicConfig(filename='error.log', level=logging.ERROR)
    now = datetime.now()
    logging.error('SystemError:'+now.strftime("%m/%d/%Y, %H:%M:%S")+str(error))

def save_csvfile(results, headers, excel, instance):
    #open to update
    wb = load_workbook(excel)
    filename = "shipment.xlsx"
    destination = "outfiles/" + str(instance) + '/' + filename
    active_sheet = wb[wb.sheetnames[0]]
    max_column = 33
    line_number = 2
    #header
    active_sheet['Y1']='DeliveryID'
    active_sheet['Z1']='Tracking Number'
    active_sheet['AA1']='Link'
    for row in active_sheet.iter_rows(line_number, active_sheet.max_row, 1, max_column):
        if row[0].value is None:
            continue
        #look for data
        for res in results:
            if str(row[0].value) == str(res[1]):
                if (res[3] is not None):
                    row[24].value = res[3]
                    row[26].value = '=HYPERLINK("{}.pdf","Label")'.format(row[0].value)
                    #copy pdf if not exist
                    if not os.path.isfile("outfiles/" + str(instance) + '/' + res[1] + '.pdf'):
                        url='https://secure.4log.com/archive/documents/' + res[4] + '.pdf'
                        r = requests.get(url, stream=True)
                        with open("outfiles/" + str(instance) + '/' + res[1] + '.pdf', 'wb') as handle:
                            for data in r.iter_content():
                                handle.write(data)
                else:
                    row[24].value = ""
                    row[26].value = res[2]
                if (res[4] is not None):
                    row[25].value = "'{}".format(res[4])
                else:
                    row[25].value = ""
    wb.save(destination)

def upload_orders(handler, instance):
    if len(handler.requests)==0:
        print("no instance requests")
    answer=input ("send to xpi?[y/n]")
    while answer!='y' and answer !='n':
        answer=input ("send to xpi?[y/n]")
    if answer=='n':
        return None
    print("sending")
    out_folder=handler.conf[4]
    results=[]
    for idx,request in enumerate(handler.requests):     
        result=create_delivery(request,out_folder)
        results.append(list(result.values()))
        if idx==0:
            headers=list(result.keys())
    print(tabulate.tabulate(results,headers=headers))
    save_csvfile(results, headers, handler.conf[6], instance)

def upload_shipments(handler, instance):
    if len(handler.requests)==0:
        print("no instance requests")
    answers=input ("send to xpi?[y/n]")
    while answers!='y' and answers !='n':
        answers=input ("send to xpi?[y/n]")
    if answers=='n':
        return None
    print("sending")
    out_folder=handler.conf[4]
    results=[]
    ordernumber=""
    awb=""
    total=len(handler.requests)
    print('Total:',total)
    with alive_bar(total) as bar:
        for idx,request in enumerate(handler.requests):
            result=create_delivery(request,out_folder)
            #get the deliveryid
            deliveryID = result['Delivery Id']
            status = result['Status']
            #validate SO already exist
            if (status.find('Header already exist for SO')>0):
                root_key=list(request.keys())[0]
                ordernumber = request[root_key]['s']['OrderNumber']
                username = request[root_key]['userName']
                deliveryID,awb = get_delivery_details(ordernumber,username).split(',')
                if (awb==''):
                    deliveryID=int(deliveryID)
            if (isinstance(deliveryID, int)):
                root_key=list(request.keys())[0]
                #add delivery to request
                request[root_key]['s']['Fourlogref']=int(deliveryID)
                request['CreateShipment'] = request.pop('CreateDelivery')
                #create the shipment
                result2=create_delivery(request,out_folder)
                #save values
                results.append(list(result2.values()))
            else:
                #update delivery and tracking
                result['Delivery Id']=deliveryID
                result['Tracking']=awb
                #save values
                results.append(list(result.values()))
            if idx==0:
                headers=list(result.keys())
            bar()
    print(tabulate.tabulate(results,headers=headers))
    save_csvfile(results, headers, handler.conf[6], instance)
    save_pdffile('outfiles/' + str(instance) + '/')
	
orders=[]
excel_handler=None
actions = {
        '1': "Convert Excel",
        '2': "Upload Orders",
		'3': "Upload Shipments",
        '4': "Quit"
    }

while True:
    try:
        selectedaction = build_text_menu(actions,'action','Main')
    
        if selectedaction == '1':
            excel_handler= Excel_Handler(0)
            excel_handler.convert_lines()
        elif selectedaction == '2' or selectedaction == '3':
            if excel_handler is not None:
                message="Instance["+str(excel_handler.instance)+"]:"
                instance=input(message)
                if instance=="":
                    instance=excel_handler.instance
            else:
                instance=input("Instance:")     
                while instance=="":
                    instance=input("Instance:")                    
            if excel_handler is None or instance!=excel_handler.instance:
                excel_handler=Excel_Handler(instance)
            if len(excel_handler.requests)>0:
                if selectedaction == '2':
                    upload_orders(excel_handler, instance)
                if selectedaction == '3':
                    upload_shipments(excel_handler, instance)
        elif selectedaction == '4':
            sys.exit() 
        else:
            errhandler(selectedaction)
            continue
    except KeyboardInterrupt:
        sys.exit()
    except Exception as e:
        log(sys.exc_info()[0])
        log(sys.exc_info()[1])
        log(sys.exc_info()[2])
        print("An exception ocurred. Verify with administrator.")