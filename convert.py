import csv
import copy
import json
import os
import shutil
import openpyxl as exl
import easygui
import time 
import random
import sys
#import colorama
from pathlib import Path
from tabulate import tabulate
from utils  import build_text_menu
#from colorama import Fore, Style

class Excel_Handler:
    def __init__(self,instance_num):
        if instance_num==0:
            self.conf=self.__init_conf(0)
            if self.conf is None:
                print('problem initializing')
                exit()
            self.requests=[]
        else:
            self.conf=self.__init_conf(instance_num)
            self.requests=self.load_instance()

    def __init_folder(self,rand):
        script_dir = Path(os.path.dirname(__file__)) 
        folder_name = 'outfiles/'+str(rand)+'/'
        working_folder = os.path.dirname(folder_name)
        try:
            os.makedirs(os.path.join(script_dir,working_folder), exist_ok=False)
        except FileExistsError as e:
            try:
                shutil.rmtree(working_folder)
                os.makedirs(working_folder, exist_ok=False)
            except Exception as e:
                print("error initializing:"+str(e))
                return None
        return os.path.join(script_dir,working_folder)

    def __init_conf(self,instance_num):
        rand = 0
        if instance_num==0:            
            rand= random.randint(1, 1000000)
            print("instance "+str(rand))
            working_folder=self.__init_folder(rand)  
        else:
            working_folder= os.path.dirname('outfiles/'+str(instance_num)+'/')
            print("loading instance "+str(instance_num))
        self.instance=rand
        resource_path=os.path.join(Path(os.path.dirname(__file__)),'resources')    

        try:    
            with open(os.path.join(resource_path,'country-codes.txt')) as csvfile:
                reader = csv.reader(csvfile, delimiter='\t')
                country_map = {rows[0].upper(): rows[1].upper() for rows in reader}
        except Exception as e:
            print("missing country file country-codes.txt",e)
            # print(e)
            return None
        try:
            with open(os.path.join(resource_path,'delivery-request-schema.json')) as json_file:
                schema= json.load(json_file)
        except:
            print("missing schema file delivery-request-schema.json")
            return None
        try:
            with open(os.path.join(resource_path,'shipment-request-schema.json')) as json_file:
                shipment_schema= json.load(json_file)
        except:
            print("missing schema file shipment-request-schema.json")
            return None
        fields_map = {'OrderNumber':0,
                'FirstName':1,
                'LastName':2,
                'CompanyName':3,               
                'AddressLine1':4,
                'AddressLine2':5,
                'AddressLine3City':6,
                'AddressLine4State':7,
                'Postcode':8, 
                'CountryCode':9,
                'EndContactEmail':10,
                'EndContactPhone':11,
                'ShipmentItemsArray.ProductHarmonizedCode':12,
                'ShipmentItemsArray.ProductDescription':13,
                'ShipmentItemsArray.ProductUnitWeight':14, 
                'ShipmentItemsArray.ProductUnitValue':15,
                'ShipmentItemsArray.ProductQuantity':16,
                'ShipmentItemsArray.ProductItemOrigin':17,
                'Currency':18,
                'ParcelWeight':19,
                'Length':21,
                'Width':22,
                'Height':23,
                'Fourlogref':24,
                'ShippingTerms':25,
                'InvoiceNumber':26,
                'Vat':27,
                'PurchasedBy':28,
                'Uom':29,
                'UomDim':30}
        if instance_num==0:
            worksheet, original= self.load_excel()
        else:
            worksheet=None
        return [country_map,schema,shipment_schema,fields_map,working_folder,worksheet,original]

    def fix_request(self,schema,request):
        root_key=list(request.keys())[0]
        request[root_key]['s']['ParcelWeight']=request[root_key]['s']['ShipmentItemsArray'][0]['ProductUnitWeight']

        if request[root_key]['s']['LastName'] is None or  request[root_key]['s']['LastName']=="":
            first_name=request[root_key]['s']['FirstName'].strip()
            if len(first_name.split(' '))>1:
                request[root_key]['s']['LastName']=request[root_key]['s']['FirstName'].split(' ')[-1]
                request[root_key]['s']['FirstName']=' '.join(first_name.split(' ')[:-1])

        if request[root_key]['s']['CountryCode']=='US':
            request[root_key]['s']['ShipType']=1
        else:
            request[root_key]['s']['ShipType']=2

        for value in request[root_key]['s'].items():
            #print("Value",value,"Type",type(value))
            if isinstance(value,tuple):
                if value[0]=="AddressLine1" or value[0]=="AddressLine2" or value[0]=="AddressLine3City" or value[0]=="EndContactEmail" or value[0]=="FirstName":
                    try:
                        ''.join(value[1]).encode('ascii')
                    except:
                        #print(f"{Fore.YELLOW}","warning - non ASCII characters for order ",request[root_key]['s']['OrderNumber'],"value:",value,f"{Style.RESET_ALL}")
                        print("warning - non ASCII characters for order ",request[root_key]['s']['OrderNumber'],"value:",value)
                    else:
                        if ''.join(value[1]).find('\\u')>0:
                            #print(f"{Fore.YELLOW}","warning - characters \\u for order ",request[root_key]['s']['OrderNumber'],"value:",value,f"{Style.RESET_ALL}")
                            print("warning - characters \\u for order ",request[root_key]['s']['OrderNumber'],"value:",value)

    def line_to_json(self,schema,line,country_map,fields_map,vendor,env,user):
        root_key=list(schema.keys())[0]
        delivery_request=copy.deepcopy(schema)
        delivery_request[root_key]["userName"]=user.username
        delivery_request[root_key]["password"]=user.password
        for schema_key,data_location in fields_map.items():
            input_val=line[data_location].value
            
            if isinstance(input_val, str):
                input_val=line[data_location].value
                input_val=input_val.strip()
                
            if '.' not in schema_key:
                #Verify that the field is mandatory and contain a value.
                if schema_key == 'EndContactPhone' and input_val is None:
                    print("The field Phone is mandatory for order ",delivery_request[root_key]['s']['OrderNumber'],"Column:",schema_key,"Value:",str(input_val))
                    return None

                if schema[root_key]['s'][schema_key]==-1:
                    if input_val is None:
                        delivery_request[root_key]['s'][schema_key]=0
                    else:
                        if isinstance(input_val, str):
                            tmp=""
                            for s in input_val:
                                if s.isdigit() or s=='.':                          
                                    tmp=tmp+s
                            try:
                                numeric_value=float(tmp)
                            except:
                                #print("Position",self.Position(fields_map, schema_key))
                                print("Please add a correct value for order ",delivery_request[root_key]['s']['OrderNumber'],"Column:",schema_key,"Value:",tmp)
                                return  None
                        else:
                            numeric_value=input_val
                        delivery_request[root_key]['s'][schema_key]=numeric_value
                else:
                    if schema_key == 'CountryCode':
                        country_name = str(input_val).upper()
                        if country_name is not None and country_name in country_map.keys():
                            country_code = country_map[country_name]
                            delivery_request[root_key]['s']['CountryCode']=country_code
                        elif country_name is not None and country_name in country_map.values():
                            country_code = country_name
                            delivery_request[root_key]['s']['CountryCode']=country_code
                        else:
                            print("No match for country:",str(country_name)," for order ",delivery_request[root_key]['s']['OrderNumber'],"Column:",schema_key,"Value:",country_name)
                            return  None

                        #The Country US have to include the State.
                        if country_name == 'UNITED STATES' or country_name == 'US':
                            state_name = delivery_request[root_key]['s']['AddressLine4State']
                            if state_name is None or state_name.__len__()==0:
                                print("The state is mandatory for the country United States"," for order ",delivery_request[root_key]['s']['OrderNumber'],"Column:",schema_key,"Value: None")
                                return None
                            else:
                                if state_name.__len__() > 2:
                                    delivery_request[root_key]['s']['AddressLine4State']=self.ConvertState(state_name)

                    else:
                        if input_val is None:
                            delivery_request[root_key]['s'][schema_key]=""
                        else:
                            delivery_request[root_key]['s'][schema_key]=str(input_val)
            else:
                node_key=schema_key.split('.')[1]
                if schema[root_key]['s']['ShipmentItemsArray'][0][node_key]==-1:
                    if input_val is None:
                        delivery_request[root_key]['s']['ShipmentItemsArray'][0][node_key]=0
                    else:
                        if isinstance(input_val, str):
                            tmp=""
                            for s in input_val:
                                if s.isdigit() or s=='.':                          
                                    tmp=tmp+s
                            numeric_value=float(tmp)
                        else:
                            numeric_value=input_val
                        delivery_request[root_key]['s']['ShipmentItemsArray'][0][node_key]=numeric_value
                else:
                    if input_val is None:
                        delivery_request[root_key]['s'][schema_key]=""
                    else:
                        delivery_request[root_key]['s']['ShipmentItemsArray'][0][node_key]=str(input_val)
        return delivery_request

    def ConvertState(self, state):
        states = {"Alabama":"AL",
                "Alaska":"AK",
                "Arizona":"AZ",
                "Arkansas":"AR",
                "California":"CA",
                "Colorado":"CO",
                "Connecticut":"CT",
                "Delaware":"DE",
                "Florida":"FL",
                "Georgia":"GA",
                "Hawaii":"HI",
                "Idaho":"ID",
                "Illinois":"IL",
                "Indiana":"IN",
                "Iowa":"IA",
                "Kansas":"KS",
                "Kentucky":"KY",
                "Louisiana":"LA",
                "Maine":"ME",
                "Maryland":"MD",
                "Massachusetts":"MA",
                "Michigan":"MI",
                "Minnesota":"MN",
                "Mississippi":"MS",
                "Missouri":"MO",
                "Montana":"MT",
                "Nebraska":"NE",
                "Nevada":"NV",
                "New Hampshire":"NH",
                "New Jersey":"NJ",
                "New Mexico":"NM",
                "New York":"NY",
                "North Carolina":"NC",
                "North Dakota":"ND",
                "Ohio":"OH",
                "Oklahoma":"OK",
                "Oregon":"OR",
                "Pennsylvania":"PA",
                "Rhode Island":"RI",
                "South Carolina":"SC",
                "South Dakota":"SD",
                "Tennessee":"TN",
                "Texas":"TX",
                "Utah":"UT",
                "Vermont":"VT",
                "Virginia":"VA",
                "Washington":"WA",
                "West Virginia":"WV",
                "Wisconsin":"WI",
                "Wyoming":"WY",
                "District of Columbia":"DC"}
        stateConverted=state
        search=states.get(state)
        if (search is not None):
            if (search.__len__() >0):
                stateConverted=search
        return stateConverted

    def load_excel(self):
        orders_file_name = easygui.fileopenbox(msg="select orders file to upload",)
        if orders_file_name is None:
            return None
        extension = os.path.splitext(orders_file_name)[1]
        try:
            if extension == '.csv':
                ordersfilename = os.path.join(os.path.splitext(orders_file_name)[0]+'.xlsx')
                wbk = exl.Workbook()
                sh = wbk.active
                with open(orders_file_name, 'r') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        row[8]=row[8].replace('\"','').replace('=','')
                        sh.append(row)
                wbk.save(ordersfilename)
                wb = exl.load_workbook(filename=ordersfilename, data_only=True)
                orders_file_name = ordersfilename
                #os.remove(ordersfilename)
            else:
                wb = exl.load_workbook(filename=orders_file_name, data_only=True)
            active_sheet = wb[wb.sheetnames[0]]
        except Exception as e:
            print("Problem reading excel file:"+str(e))
            return None
        return active_sheet, orders_file_name

    def convert_lines(self):
        if len(self.requests)>0:
            return self.requests
        converted=[]
        for e in self.conf:
            if e is None:
                print("no conf")
                return None
        conf=self.conf
        out_folder=conf[4]
        delivery_schema=conf[1]
        shipment_schema=conf[2]
        country_map=conf[0]
        fields_map=conf[3]
        active_sheet = conf[5]
        vendor=input("vendor:[JWG]")
        if vendor=="":
            vendor="JWG"
        env=input("environment:[p]")       
        if env=="":
            env="p"
        max_column = 31
        line_number = 2
        #verify profile of the user
        prof = profile()
        user = prof.manageprofile(env, None, vendor)
        if user is None or user.username=="":
            return None
        print("converting")
        for row in active_sheet.iter_rows(2, active_sheet.max_row, 1, max_column):
            if row[0].value is None:
                continue
            #if 4log refernce exist
            print("4LogReference:",str(row[24].value),"OrderNumber:",str(row[0].value))
            if row[24].value is not None:
                active_schema=shipment_schema
            else:
                active_schema=delivery_schema
            so_number=row[0].value
            converted_request=self.line_to_json(active_schema,row,country_map,fields_map,vendor,env,user)
            if converted_request is not None:
                self.fix_request(active_schema,converted_request)
                with open(os.path.join(out_folder,str(so_number)+'_request.json'),"w",) as request_file:
                    request_file.write(json.dumps(converted_request,indent=2))
                    converted.append(converted_request)
            else:
                print("skipping line number "+str(line_number))
            line_number=line_number+1
        self.requests=converted
        return(converted)

    def load_instance(self):
        folder=self.conf[4]
        requests=[]
        try:
            for file in enumerate(os.listdir(folder)):
                if file.endswith("_request.json"):
                    requests.append(json.loads(open(os.path.join(folder,file)).read()))
        except:
            pass
        if len(requests)==0:
            print("nothing found")
        self.requests=requests
        return requests

class profile:
    pass

    #manage the profile of the user. parameters: environment, schema and client name
    def manageprofile(self,env,schema_type,client):
        resource_path=os.path.join(Path(os.path.dirname(__file__)),'resources')

        #if environment is empty, the default environment is production
        if env=="":
            env="p"
        #set the schema type: 1 - Delivery, 2 - SO
        schema="Delivery"
        try:    
            _username=""
            _password=""
            _vendor=""			  
            #look in the profile if the user match the condition to convert Excel
            with open(os.path.join(resource_path,'profile.txt')) as csvfile:
                reader = csv.reader(csvfile, delimiter='\t')
                for rows in reader:
                    if rows[0]==env:
                        if rows[1]==schema:
                            if rows[2]==client:
                                _username=rows[3]
                                _password=rows[4]
                            if rows[3]==client:
                                _vendor=rows[5]
            #if no match, send a message to the user in the screen
            if _username=="" and _vendor is None:
                print("No valid configuration for this user, schema and environment")
                return None
            #profile matched
            result = profile()
            result.username = _username
            result.password = _password
            result.vendor = _vendor
            return result

        except Exception as e:
            print("missing profile file profile.txt",e)       
            return None